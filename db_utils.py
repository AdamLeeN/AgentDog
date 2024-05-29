from typing import TYPE_CHECKING, Any, Dict, List, Optional, Sequence, TypeVar, Union
from uuid import UUID
import inspect
import json

import pandas as pd
import numpy as np
import re
import os
import shutil

from langchain_community.document_loaders import TextLoader
from langchain.schema import Document
from langchain_community.embeddings.huggingface import HuggingFaceEmbeddings
from langchain_community.vectorstores import Chroma
from langchain_text_splitters import CharacterTextSplitter
from openpyxl import load_workbook



def load_embedding_model():
    """
    加载embedding模型
    :return: HuggingFaceEmbeddings
    """
    encode_kwargs = {"normalize_embeddings": False} #指定在进行编码（embedding）时是否对向量进行归一化
    model_kwargs = {"device": "cuda:0"}
    return HuggingFaceEmbeddings(
        model_name="model/models--BAAI--bge-large-zh-v1.5/snapshots/79e7739b6ab944e86d6171e44d24c997fc1e0116",
        model_kwargs=model_kwargs,
        encode_kwargs=encode_kwargs
    )




def load_db(embedding, persist_directory='VectorStore'):
    """
    加载数据库
    """
    if not os.path.exists(persist_directory):
        loader = TextLoader("know2.txt", encoding='UTF-8')
        documents = loader.load()
        text_splitter = CharacterTextSplitter(chunk_size=16, chunk_overlap=0)
        docs = text_splitter.split_documents(documents)
        db = Chroma.from_documents(docs, embedding, persist_directory=persist_directory)
        db.persist()
    else: # 从已有数据中加载
        db = Chroma(persist_directory=persist_directory, embedding_function=embedding)
    return db





def add_record(db, data: str, thred: float=0.8, k: int=1):
    """
    添加新的记录，作为后续few-shot的示例。
    :param db: 数据库对象
    :param data: 需要添加的数据，格式为：状态：xxx\n动作：xxx
    :param thred: 余弦形似度大于这个数值的进行判断
    :return: None
    """
    
    for simi_data, socre in db.similarity_search_with_score(data, k=k):
        if(socre < 1-thred):
            #score较小，相似
            percentage = (1-socre)*100
            if(int(percentage)==100):
                break

            print('-------------------------------')
            print(f'已有数据:  {simi_data.page_content}')
            print(f'新的数据:  {data}')
            user_input = input(f'他们的相似度为{percentage:.3f}%，是否继续添加数据？(yes or no)').strip().lower()
            if(user_input=='yes'):
                db.add_documents([Document(page_content=data)])
                print(f'添加成功！')
                break
            else:
                print(f'跳过！')
                break
        else:
            #score较大，不够相似
            print('-------------------------------')
            print(f'新的数据:  {data}')
            db.add_documents([Document(page_content=data)])
            print(f'添加成功！')
            break



def read_data(X, Y, s, file='data/测试案例.xlsm'):
    # 指定文件路径，data_only属性作用是取值，不取公式
    workbook = load_workbook(file, data_only=True)
    # 指定sheet页
    sheet = workbook[s]
    for i in range(2, sheet.max_row):
        result1 = sheet[f"{X}{i}"].value
        result2 = sheet[f"{Y}{i}"].value
        if(result1):
            result1 = result1.strip().split('\n')
            result2 = result2.strip().split('\n')
            merged_text1 = ';'.join(result1)
            merged_text2 = ';'.join(result2)
            data = f'状态：{merged_text1}\n指令：{merged_text2}'
            yield data
    workbook.close()


if __name__ == "__main__":
    embedding = load_embedding_model()
    #Initial_Condition_Data  Action_Data  Expected_Result_Data
    db = load_db(embedding, persist_directory='baseKnowledge_2')
    # #FI  GJ   HK
    # ds = read_data(X='H', Y='K', s='1DATM', file=r'data/12.xlsm')

    # for data in ds:
    #     add_record(db, data)
    
